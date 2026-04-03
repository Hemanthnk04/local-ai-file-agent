"""
content_validator.py — strict per-extension content validation and sanitization.

Every file written by the agent passes through here BEFORE being saved.
Purpose: ensure the model output contains ONLY valid file content,
         with no prose explanations, markdown fences, or preamble mixed in.
"""

import re
import os
import json
from io import StringIO


# ══════════════════════════════════════════════════════════════════════════════
# SANITIZERS — strip common LLM pollution from output
# ══════════════════════════════════════════════════════════════════════════════

def _strip_markdown_fences(text):
    """Remove ```lang ... ``` wrappers the LLM sometimes adds."""
    text = re.sub(r"^```[\w]*\r?\n?", "", text.lstrip(), flags=re.M)
    text = re.sub(r"\n?```\s*$", "", text.rstrip(), flags=re.M)
    return text.strip()


def _strip_preamble(text, ext):
    """
    Remove common LLM preamble lines that appear before real content.
    E.g. "Here is the Python code:", "Sure, here's the file:", etc.
    """
    preamble_patterns = [
        r"^(here is|here's|below is|the following is|i've (created|generated|written)).*:?\s*$",
        r"^(sure[,!]?|certainly[,!]?|of course[,!]?).*$",
        r"^(this (file|code|script|data|content) (contains|shows|is|has)).*$",
        r"^---+\s*$",
        r"^note:.*$",
        r"^explanation:.*$",
    ]
    lines = text.splitlines()
    start = 0
    for i, line in enumerate(lines[:15]):  # check first 15 lines for preamble
        stripped = line.strip().lower()
        if any(re.match(p, stripped, re.I) for p in preamble_patterns):
            start = i + 1
        else:
            break
    return "\n".join(lines[start:]).strip()


def _strip_postamble(text):
    """Remove common LLM sign-off lines at the end."""
    postamble_patterns = [
        r"^(this (code|script|file|data) (will|should|can|does)).*$",
        r"^(let me know if).*$",
        r"^(feel free to).*$",
        r"^(you can (run|use|modify|adjust)).*$",
        r"^(note:).*$",
        r"^(i hope this helps).*$",
    ]
    lines = text.splitlines()
    end = len(lines)
    for i in range(len(lines) - 1, max(len(lines) - 6, -1), -1):
        stripped = lines[i].strip().lower()
        if any(re.match(p, stripped, re.I) for p in postamble_patterns):
            end = i
        else:
            break
    return "\n".join(lines[:end]).rstrip()


def sanitize(text, ext):
    """
    Full sanitization pipeline for a given file extension.
    Returns cleaned text.
    """
    if not text or not text.strip():
        return text

    text = _strip_markdown_fences(text)
    text = _strip_preamble(text, ext)
    text = _strip_postamble(text)
    return text


# ══════════════════════════════════════════════════════════════════════════════
# STRICT VALIDATORS — return (ok, reason, fixed_content)
# fixed_content is the sanitized version if fixable, else original
# ══════════════════════════════════════════════════════════════════════════════

def _validate_csv(content):
    try:
        import pandas as pd
        df = pd.read_csv(StringIO(content))
        if df.empty:
            return False, "CSV is empty — no data rows", content
        if len(df.columns) == 1 and df.columns[0].strip().lower().startswith(("here", "sure", "below")):
            return False, "CSV appears to contain LLM prose, not data", content
        return True, f"Valid CSV: {len(df)} rows × {len(df.columns)} columns", content
    except Exception as e:
        return False, f"Invalid CSV structure: {e}", content



def _validate_tsv(content):
    try:
        import pandas as pd
        df = pd.read_csv(StringIO(content), sep="\t")
        if df.empty:
            return False, "TSV is empty — no data rows", content
        return True, f"Valid TSV: {len(df)} rows × {len(df.columns)} columns", content
    except Exception as e:
        return False, f"Invalid TSV structure: {e}", content


def _validate_json(content):
    try:
        data = json.loads(content)
        if isinstance(data, str) and len(data) > 50:
            return False, "JSON is a plain string — likely LLM prose", content
        kind = "array" if isinstance(data, list) else type(data).__name__
        count = len(data) if isinstance(data, (list, dict)) else 1
        return True, f"Valid JSON: {kind} with {count} items", content
    except json.JSONDecodeError as e:
        return False, f"Invalid JSON: {e}", content


def _validate_xml(content):
    try:
        import xml.etree.ElementTree as ET
        ET.fromstring(content)
        return True, "Valid XML", content
    except Exception as e:
        return False, f"Invalid XML: {e}", content


def _validate_yaml(content):
    try:
        import yaml
        data = yaml.safe_load(content)
        if isinstance(data, str) and "\n" not in content:
            return False, "YAML looks like plain text, not structured data", content
        return True, f"Valid YAML ({type(data).__name__})", content
    except Exception as e:
        return False, f"Invalid YAML: {e}", content


def _validate_python(content):
    import ast
    try:
        ast.parse(content)
        # Check for prose contamination: non-code text before first statement
        lines = content.splitlines()
        for line in lines[:5]:
            stripped = line.strip()
            if stripped and not stripped.startswith("#") and not stripped.startswith("\"\"\""):
                if re.match(r"^(here|sure|below|this code|the following)", stripped, re.I):
                    return False, "Python file starts with LLM prose, not code", content
        return True, "Valid Python", content
    except SyntaxError as e:
        return False, f"Python SyntaxError at line {e.lineno}: {e.msg}", content


def _validate_html(content):
    # Must contain actual HTML tags
    if not re.search(r"<[a-zA-Z][^>]*>", content):
        return False, "No HTML tags found — content may be plain text", content
    # Check basic structure
    has_doctype = "<!doctype" in content.lower() or "<html" in content.lower()
    if not has_doctype:
        return False, "Missing DOCTYPE or <html> tag", content
    return True, "Valid HTML", content


def _validate_js_ts(content, lang="JavaScript"):
    # Check for obvious prose
    lines = [l.strip() for l in content.splitlines() if l.strip()]
    if lines and re.match(r"^(here|sure|below|this|the)", lines[0], re.I):
        return False, f"{lang} file starts with prose", content
    # Basic brace balance
    opens  = content.count("{")
    closes = content.count("}")
    if abs(opens - closes) > 2:
        return False, f"{lang} brace mismatch: {opens}{{ vs {closes}}}", content
    return True, f"Valid {lang}", content


def _validate_java(content):
    if not re.search(r"\bclass\b|\binterface\b|\benum\b", content):
        return False, "No class/interface/enum declaration found", content
    opens  = content.count("{")
    closes = content.count("}")
    if abs(opens - closes) > 2:
        return False, f"Java brace mismatch: {opens}{{ vs {closes}}}", content
    return True, "Valid Java", content


def _validate_sql(content):
    keywords = re.findall(
        r"\b(SELECT|INSERT|UPDATE|DELETE|CREATE|DROP|ALTER|WITH|FROM|WHERE)\b",
        content, re.I
    )
    if not keywords:
        return False, "No SQL keywords found", content
    return True, f"Valid SQL — keywords: {', '.join(sorted(set(k.upper() for k in keywords)))}", content


def _validate_generic_code(content, lang):
    """Generic check for C, C++, C#, Go, Ruby, PHP etc."""
    if not content.strip():
        return False, f"Empty {lang} file", content
    lines = [l.strip() for l in content.splitlines() if l.strip()]
    if lines and re.match(r"^(here|sure|below|this|the following)", lines[0], re.I):
        return False, f"{lang} file starts with LLM prose", content
    return True, f"Valid {lang} (structure check passed)", content


def _validate_md(content):
    if not content.strip():
        return False, "Empty Markdown file", content
    return True, f"Valid Markdown ({len(content.splitlines())} lines)", content


def _validate_toml(content):
    try:
        try:
            import tomllib
            tomllib.loads(content)
        except ImportError:
            import tomli
            tomli.loads(content)
        return True, "Valid TOML", content
    except ImportError:
        return True, "TOML check skipped (tomllib not installed)", content
    except Exception as e:
        return False, f"Invalid TOML: {e}", content


def _validate_ini(content):
    import configparser
    try:
        cfg = configparser.ConfigParser()
        cfg.read_string(content)
        return True, f"Valid INI — {len(cfg.sections())} sections", content
    except Exception as e:
        return False, f"Invalid INI: {e}", content


# Extension → validator function
_EXT_VALIDATORS = {
    ".csv":      _validate_csv,
    ".tsv":      _validate_tsv,
    ".json":     _validate_json,
    ".xml":      _validate_xml,
    ".yaml":     _validate_yaml,
    ".yml":      _validate_yaml,
    ".py":       _validate_python,
    ".pyw":      _validate_python,
    ".html":     _validate_html,
    ".htm":      _validate_html,
    ".js":       lambda c: _validate_js_ts(c, "JavaScript"),
    ".ts":       lambda c: _validate_js_ts(c, "TypeScript"),
    ".jsx":      lambda c: _validate_js_ts(c, "JavaScript"),
    ".tsx":      lambda c: _validate_js_ts(c, "TypeScript"),
    ".java":     _validate_java,
    ".sql":      _validate_sql,
    ".md":       _validate_md,
    ".markdown": _validate_md,
    ".toml":     _validate_toml,
    ".ini":      _validate_ini,
    ".cfg":      _validate_ini,
    ".cs":       lambda c: _validate_generic_code(c, "C#"),
    ".cpp":      lambda c: _validate_generic_code(c, "C++"),
    ".c":        lambda c: _validate_generic_code(c, "C"),
    ".go":       lambda c: _validate_generic_code(c, "Go"),
    ".rb":       lambda c: _validate_generic_code(c, "Ruby"),
    ".php":      lambda c: _validate_generic_code(c, "PHP"),
    ".css":      lambda c: _validate_generic_code(c, "CSS"),
}


def validate_content(content, filename):
    """
    Sanitize then strictly validate content for the given filename.
    Returns (ok, reason, clean_content).
    """
    if not content:
        return False, "Empty content", content

    ext = os.path.splitext(filename)[1].lower() if filename else ""

    # Step 1: sanitize
    clean = sanitize(content, ext)

    # Step 2: validate
    validator = _EXT_VALIDATORS.get(ext)
    if validator:
        ok, reason, clean = validator(clean)
        return ok, reason, clean

    # No specific validator — just ensure it's not pure LLM prose
    lines = [l.strip() for l in clean.splitlines() if l.strip()]
    if lines and re.match(r"^(here is|here's|sure[,!]|certainly)", lines[0], re.I):
        return False, "Content appears to be LLM explanation, not file data", clean

    return True, f"Content accepted ({ext or 'unknown type'})", clean


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL MULTI-SHEET BUILDER
# ══════════════════════════════════════════════════════════════════════════════

def build_excel_multisheet(sheet_data, out_path):
    """
    Build a proper multi-sheet Excel file.

    sheet_data: dict of {sheet_name: content}
    content can be:
      - pd.DataFrame
      - str (CSV text)
      - list of lists ([[header,...], [row,...]])

    Returns (ok, reason).
    """
    import pandas as pd
    from openpyxl import Workbook

    wb     = Workbook()
    first  = True
    errors = []

    for sheet_name, content in sheet_data.items():
        # Sanitize sheet name: max 31 chars, no illegal chars
        safe_name = re.sub(r"[\\/*?:\[\]]", "_", str(sheet_name))[:31]

        if first:
            ws    = wb.active
            ws.title = safe_name
            first = False
        else:
            ws = wb.create_sheet(safe_name)

        # Convert content to rows
        try:
            if isinstance(content, pd.DataFrame):
                df = content
            elif isinstance(content, str) and content.strip():
                try:
                    df = pd.read_csv(StringIO(content))
                except Exception:
                    # treat as plain text — one column
                    lines = content.splitlines()
                    df    = pd.DataFrame({"content": lines})
            elif isinstance(content, list):
                if content and isinstance(content[0], (list, tuple)):
                    df = pd.DataFrame(content[1:], columns=content[0]) if len(content) > 1 else pd.DataFrame(content)
                else:
                    df = pd.DataFrame({"value": content})
            else:
                df = pd.DataFrame({"value": [str(content)]})

            # Write header
            for col_idx, col_name in enumerate(df.columns, 1):
                ws.cell(row=1, column=col_idx, value=str(col_name))

            # Write data rows
            for row_idx, row in enumerate(df.itertuples(index=False), 2):
                for col_idx, val in enumerate(row, 1):
                    ws.cell(row=row_idx, column=col_idx, value=val)

        except Exception as e:
            errors.append(f"Sheet '{safe_name}': {e}")
            ws.cell(row=1, column=1, value=f"Error: {e}")

    try:
        wb.save(out_path)
    except Exception as e:
        return False, f"Failed to save Excel: {e}"

    if errors:
        return True, f"Saved with warnings: {'; '.join(errors)}"
    return True, f"Saved {len(sheet_data)} sheet(s)"
